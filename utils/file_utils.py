import sys
import os
from pathlib import Path

def resource_path(relative_path: str) -> str:
    """
    Get absolute path to resource, works for dev and for PyInstaller.
    """
    # PyInstaller bundled runtime path
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)

    # dev environment
    return os.path.join(os.path.abspath("."), relative_path)


def load_prompt(filename: str) -> str:
    """
    Load the content of a prompt text file and return it as a string.
    Automatically supports PyInstaller bundled files.
    """
    real_path = resource_path(filename)
    prompt_path = Path(real_path)

    if not prompt_path.exists():
        raise FileNotFoundError(f"Prompt file not found: {prompt_path}")

    with prompt_path.open("r", encoding="utf-8") as f:
        return f.read()


import pandas as pd
from openpyxl import load_workbook

def get_hyperlinks_from_excel(file_path: str, column_name: str = "Link") -> list:
    """
    从指定的 Excel 文件中提取指定列（默认是 'Link'）的超链接。

    :param file_path: Excel 文件路径
    :param column_name: 要查找的列名（默认为 'Link'）
    :return: 包含超链接的列表
    """
    # 读取 Excel 文件并获取 DataFrame
    df = pd.read_excel(file_path, engine='openpyxl')

    # 加载 Excel 文件并获取工作表
    wb = load_workbook(file_path)
    ws = wb.active  # 或者指定具体的工作表，如 wb['Sheet1']

    # 查找 Link 列的列名
    link_column = None
    for col in df.columns:
        if col.lower() == column_name.lower():  # 忽略大小写匹配列名为 "Link"
            link_column = col
            break

    if not link_column:
        raise ValueError(f"未找到名为 '{column_name}' 的列")

    # 获取 Link 列的所有超链接
    links = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # 从第二行开始（假设第一行是表头）
        for cell in row:
            if cell.column == df.columns.get_loc(link_column) + 1:  # 确保是 Link 列
                if cell.hyperlink:
                    links.append(cell.hyperlink.target)

    return links

